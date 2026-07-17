from __future__ import annotations

import io
import json
import os
import re
import shutil
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_EXPORT_LIMIT = 500
SET_TITLE = "Разом дешевше"
PUBLIC_LOG_PATH_JSON_PATTERN = re.compile(r'("public_log_path"\s*:\s*")(?P<value>[^"]*)(")')


class HoroshopSetsError(RuntimeError):
    pass


@dataclass(frozen=True)
class Settings:
    domain: str
    host: str
    port: int
    currency: str
    title: str
    batch_size: int
    request_timeout_seconds: int
    state_file: Path
    public_log_path: Path = Path("logs")
    public_log_name: str = "horoshop_sets.log"

    @property
    def public_log_file(self) -> Path:
        return self.public_log_path / self.public_log_name


@dataclass(frozen=True)
class Credentials:
    login: str
    password: str
    token: str = ""


@dataclass(frozen=True)
class SetRow:
    article: str
    display_articles: tuple[str, ...]
    discounted_price: Decimal | None
    row_number: int
    action: str = "upsert"
    title: str = ""
    enabled: bool = True
    sort_order: int | None = None
    discount_percent: int | None = None
    currency: str = ""


@dataclass(frozen=True)
class CatalogProduct:
    article: str
    article_for_display: str


@dataclass(frozen=True)
class PlanItem:
    article: str
    display_articles: tuple[str, ...]
    discounted_price: Decimal | None
    products: tuple[str, ...]
    row_number: int
    error: str = ""
    action: str = "upsert"
    title: str = ""
    enabled: bool = True
    sort_order: int | None = None
    discount_percent: int | None = None
    currency: str = ""

    @property
    def ready(self) -> bool:
        return not self.error


def utc_now() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def normalize(value: Any) -> str:
    return "" if value is None else str(value).strip()


def endpoint_url(domain: str, endpoint: str) -> str:
    return urljoin(f"{domain.rstrip('/')}/", endpoint.lstrip("/"))


def repair_public_log_path_json(config_text: str) -> str:
    def replace_path(match: re.Match[str]) -> str:
        path_value = re.sub(r"(?<!\\)\\(?!\\)", lambda _: "\\\\", match.group("value"))
        return f'{match.group(1)}{path_value}{match.group(3)}'

    return PUBLIC_LOG_PATH_JSON_PATTERN.sub(replace_path, config_text)


def load_settings(config_file: Path) -> Settings:
    config_text = config_file.read_text(encoding="utf-8-sig")
    try:
        raw = json.loads(config_text)
    except json.JSONDecodeError as original_error:
        repaired_config_text = repair_public_log_path_json(config_text)
        if repaired_config_text == config_text:
            raise original_error
        try:
            raw = json.loads(repaired_config_text)
        except json.JSONDecodeError:
            raise original_error
        try:
            config_file.write_text(repaired_config_text, encoding="utf-8")
        except OSError:
            pass
    if not isinstance(raw, dict):
        raise ValueError("config.json must contain an object.")

    horoshop = raw.get("horoshop") or {}
    server = raw.get("server") or {}
    logging_config = raw.get("logging") or {}
    if not isinstance(horoshop, dict) or not isinstance(server, dict) or not isinstance(logging_config, dict):
        raise ValueError("Sections horoshop, server and logging must contain objects.")

    domain = normalize(horoshop.get("domain"))
    if not domain:
        raise ValueError("Set horoshop.domain in config.json.")
    currency = normalize(horoshop.get("currency", "UAH")).upper()
    if len(currency) != 3:
        raise ValueError("horoshop.currency must be a three-letter ISO code.")

    state_value = normalize(server.get("state_file", "data/sets_state.json"))
    state_file = Path(state_value)
    if not state_file.is_absolute():
        state_file = config_file.parent / state_file
    public_log_path = Path(normalize(logging_config.get("public_log_path", "logs")) or "logs")
    if not public_log_path.is_absolute():
        public_log_path = config_file.parent / public_log_path
    public_log_name = normalize(logging_config.get("public_log_name", "horoshop_sets.log")) or "horoshop_sets.log"
    if Path(public_log_name).name != public_log_name:
        raise ValueError("logging.public_log_name must be a file name without a path.")
    title = SET_TITLE

    return Settings(
        domain=domain.rstrip("/"),
        host=normalize(server.get("host", "0.0.0.0")) or "0.0.0.0",
        port=max(1, min(65535, int(server.get("port", 8093)))),
        currency=currency,
        title=title,
        batch_size=max(1, int(horoshop.get("batch_size", 50))),
        request_timeout_seconds=max(
            1, int(horoshop.get("request_timeout_seconds", 60))
        ),
        state_file=state_file,
        public_log_path=public_log_path,
        public_log_name=public_log_name,
    )


def parse_price(value: Any) -> Decimal:
    text = normalize(value).replace(" ", "").replace(",", ".")
    try:
        price = Decimal(text)
    except (InvalidOperation, ValueError) as error:
        raise ValueError("ціна має бути числом, більшим за нуль.") from error
    if price <= 0:
        raise ValueError("ціна має бути більшою за нуль.")
    return price.quantize(Decimal("0.01"))


def split_display_articles(value: Any) -> tuple[str, ...]:
    values = tuple(item.strip() for item in normalize(value).split(";") if item.strip())
    if len(values) < 2:
        raise ValueError("у наборі має бути щонайменше два артикули товарів.")
    if len({item.casefold() for item in values}) != len(values):
        raise ValueError("артикули товарів у наборі не можуть повторюватися.")
    return values


def build_set_article(display_articles: tuple[str, ...]) -> str:
    return ".".join(display_articles)


def parse_action(value: Any) -> str:
    action = normalize(value).casefold()
    aliases = {
        "": "upsert",
        "обновить": "upsert",
        "создать": "upsert",
        "update": "upsert",
        "upsert": "upsert",
        "удалить": "delete",
        "delete": "delete",
        "да": "delete",
        "так": "delete",
        "yes": "delete",
        "1": "delete",
        "нет": "upsert",
        "ні": "upsert",
        "no": "upsert",
        "0": "upsert",
        "принять на учет": "register",
        "принять на учёт": "register",
        "учет": "register",
        "учёт": "register",
        "register": "register",
    }
    if action not in aliases:
        raise ValueError("у колонці «Видалити (Так)» вкажіть Так або залиште її порожньою.")
    return aliases[action]


def parse_bool(value: Any, default: bool = True) -> bool:
    text = normalize(value).casefold()
    if not text:
        return default
    if text in {"да", "true", "1", "yes", "y"}:
        return True
    if text in {"нет", "false", "0", "no", "n"}:
        return False
    raise ValueError("значение активности должно быть Да или Нет.")


def parse_optional_int(value: Any, label: str, minimum: int, maximum: int | None = None) -> int | None:
    text = normalize(value)
    if not text:
        return None
    try:
        parsed = int(text)
    except ValueError as error:
        raise ValueError(f"{label} должно быть целым числом.") from error
    if parsed < minimum or (maximum is not None and parsed > maximum):
        limits = f"от {minimum} до {maximum}" if maximum is not None else f"не меньше {minimum}"
        raise ValueError(f"{label} должно быть {limits}.")
    return parsed


def parse_excel_sets(data: bytes) -> list[SetRow]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows: list[SetRow] = []
        seen_articles: set[str] = set()
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if not row or all(value is None for value in row[:9]):
                continue
            article = normalize(row[0] if len(row) > 0 else "")
            if article.casefold() in {
                "article",
                "артикул",
                "артикул набору",
                "артикул набору (необов'язково)",
                "артикул набора",
            }:
                continue
            try:
                action = parse_action(row[3] if len(row) > 3 else "")
                if action == "delete":
                    if not article:
                        raise ValueError("для видалення вкажіть артикул набору.")
                    key = article.casefold()
                    if key in seen_articles:
                        raise ValueError(f"артикул набору '{article}' повторюється.")
                    seen_articles.add(key)
                    rows.append(SetRow(article, (), None, row_number, action=action))
                    continue
                display_articles = split_display_articles(row[1] if len(row) > 1 else "")
                article = article or build_set_article(display_articles)
                key = article.casefold()
                if key in seen_articles:
                    raise ValueError(f"артикул набору '{article}' повторюється.")
                seen_articles.add(key)
                price = parse_price(row[2] if len(row) > 2 else "")
                title = ""
                enabled = True
                sort_order = None
                discount_percent = None
                currency = ""
            except ValueError as error:
                raise ValueError(f"Рядок {row_number}: {error}") from error
            rows.append(
                SetRow(
                    article,
                    display_articles,
                    price,
                    row_number,
                    action=action,
                    title=title,
                    enabled=enabled,
                    sort_order=sort_order,
                    discount_percent=discount_percent,
                    currency=currency,
                )
            )
        if not rows:
            raise ValueError("Excel не містить жодного набору.")
        return rows
    finally:
        workbook.close()


def build_excel_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Набори"
    worksheet.append(["Артикул набору (необов'язково)", "Артикули відображення товарів", "Кінцева ціна"])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:C1"
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 58
    worksheet.column_dimensions["C"].width = 18

    header_fill = PatternFill("solid", fgColor="166534")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in range(2, 102):
        worksheet.cell(row=row, column=1).number_format = "@"
        worksheet.cell(row=row, column=2).number_format = "@"
        worksheet.cell(row=row, column=3).number_format = "0.00"

    guide = workbook.create_sheet("Інструкція")
    guide.column_dimensions["A"].width = 105
    guide["A1"] = "Шаблон для створення та оновлення наборів"
    guide["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    guide["A1"].fill = header_fill
    guide.merge_cells("A1:B1")
    instructions = [
        "Заповнюйте лист «Набори». Кожен рядок створює або оновлює один набір.",
        "Артикул набору необов'язковий: якщо залишити його порожнім, він буде складений з артикулів товарів через крапку.",
        "Артикули товарів: щонайменше два значення article_for_display, розділені крапкою з комою (;).",
        "Кінцева ціна: ціна набору після знижки.",
        "Не додавайте колонки «Дія», «Активний», «Порядок сортування», «Знижка» або «Валюта».",
        "Для масового видалення вивантажте реєстр: у ньому є окрема колонка «Видалити (Так)».",
    ]
    for row, instruction in enumerate(instructions, start=3):
        cell = guide.cell(row=row, column=1, value=instruction)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    guide.freeze_panes = "A3"

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class CatalogIndex:
    def __init__(self, products: list[CatalogProduct]):
        self.products = products
        self.product_articles = {product.article for product in products}
        self.by_display: dict[str, list[CatalogProduct]] = {}
        self.by_display_folded: dict[str, list[CatalogProduct]] = {}
        for product in products:
            if not product.article_for_display:
                continue
            self.by_display.setdefault(product.article_for_display, []).append(product)
            self.by_display_folded.setdefault(
                product.article_for_display.casefold(), []
            ).append(product)

    @classmethod
    def from_raw(cls, raw_products: list[dict[str, Any]]) -> "CatalogIndex":
        products = []
        for raw in raw_products:
            if not isinstance(raw, dict):
                continue
            article = normalize(raw.get("article"))
            if article:
                products.append(
                    CatalogProduct(
                        article=article,
                        article_for_display=normalize(raw.get("article_for_display")),
                    )
                )
        return cls(products)

    def resolve_display_article(self, display_article: str) -> tuple[str | None, str]:
        exact = self.by_display.get(display_article, [])
        if len(exact) == 1:
            return exact[0].article, ""
        if len(exact) > 1:
            return None, f"Артикул відображення '{display_article}' не є унікальним."
        insensitive = self.by_display_folded.get(display_article.casefold(), [])
        if len(insensitive) == 1:
            return insensitive[0].article, ""
        if len(insensitive) > 1:
            return None, f"Артикул відображення '{display_article}' не є унікальним."
        return None, f"Артикул відображення '{display_article}' не знайдений у каталозі."


def prepare_plan(rows: list[SetRow], catalog: CatalogIndex) -> list[PlanItem]:
    plan: list[PlanItem] = []
    for row in rows:
        if row.action == "delete":
            plan.append(
                PlanItem(
                    article=row.article,
                    display_articles=(),
                    discounted_price=None,
                    products=(),
                    row_number=row.row_number,
                    action="delete",
                )
            )
            continue
        if row.article in catalog.product_articles:
            plan.append(
                PlanItem(
                    article=row.article,
                    display_articles=row.display_articles,
                    discounted_price=row.discounted_price,
                    products=(),
                    row_number=row.row_number,
                    error="Артикул набору збігається з артикулом наявного товару.",
                    action=row.action,
                    title=row.title,
                    enabled=row.enabled,
                    sort_order=row.sort_order,
                    discount_percent=row.discount_percent,
                    currency=row.currency,
                )
            )
            continue

        products: list[str] = []
        error = ""
        for display_article in row.display_articles:
            product_article, error = catalog.resolve_display_article(display_article)
            if error:
                break
            if product_article:
                products.append(product_article)
        if not error and len(set(products)) != len(products):
            error = "Кілька артикулів відображення вказують на один товар."
        plan.append(
            PlanItem(
                article=row.article,
                display_articles=row.display_articles,
                discounted_price=row.discounted_price,
                products=tuple(products) if not error else (),
                row_number=row.row_number,
                error=error,
                action=row.action,
                title=row.title,
                enabled=row.enabled,
                sort_order=row.sort_order,
                discount_percent=row.discount_percent,
                currency=row.currency,
            )
        )
    return plan


def import_payload(items: list[PlanItem], settings: Settings) -> list[dict[str, Any]]:
    payload = []
    for item in items:
        if not item.ready or item.action != "upsert":
            continue
        value: dict[str, Any] = {
            "article": item.article,
            "title": SET_TITLE,
            "discountedPrice": float(item.discounted_price or Decimal("0")),
            "currency": item.currency or settings.currency,
            "enabled": item.enabled,
            "products": list(item.products),
        }
        payload.append(value)
    return payload


class HoroshopClient:
    def __init__(self, settings: Settings, credentials: Credentials):
        self.settings = settings
        self.credentials = credentials
        self.session = requests.Session()
        self._token = credentials.token

    def _post(self, endpoint: str, payload: dict[str, Any]) -> dict[str, Any]:
        try:
            response = self.session.post(
                endpoint_url(self.settings.domain, endpoint),
                json=payload,
                timeout=self.settings.request_timeout_seconds,
            )
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as error:
            raise HoroshopSetsError(f"Horoshop API request failed: {error}") from error
        except ValueError as error:
            raise HoroshopSetsError("Horoshop API returned a non-JSON response.") from error
        if not isinstance(data, dict):
            raise HoroshopSetsError("Horoshop API returned an invalid JSON response.")
        if str(data.get("status", "")).upper() in {"ERROR", "EXCEPTION"}:
            raise HoroshopSetsError(str(data))
        return data

    def token(self) -> str:
        if self._token:
            return self._token
        if not self.credentials.login or not self.credentials.password:
            raise ValueError("Enter Horoshop login and password.")
        response = self._post(
            "/api/auth/",
            {"login": self.credentials.login, "password": self.credentials.password},
        )
        token = response.get("response", {}).get("token")
        if not token:
            raise HoroshopSetsError("Horoshop did not return an authorization token.")
        self._token = str(token)
        return self._token

    def export_catalog(self) -> list[dict[str, Any]]:
        offset = 0
        products: list[dict[str, Any]] = []
        while True:
            response = self._post(
                "/api/catalog/export/",
                {
                    "token": self.token(),
                    "offset": offset,
                    "limit": DEFAULT_EXPORT_LIMIT,
                    "includedParams": ["article_for_display"],
                },
            )
            nested = response.get("response")
            page = nested.get("products") if isinstance(nested, dict) else response.get("products")
            if not isinstance(page, list):
                raise HoroshopSetsError("Catalog export did not contain products.")
            products.extend(item for item in page if isinstance(item, dict))
            if len(page) < DEFAULT_EXPORT_LIMIT:
                return products
            offset += DEFAULT_EXPORT_LIMIT

    def import_sets(self, items: list[dict[str, Any]]) -> dict[str, Any]:
        if not items:
            return {"status": "OK", "response": {"log": []}}
        return self._post(
            "/api/productSet/import/",
            {"token": self.token(), "items": items},
        )

    def remove_sets(self, articles: list[str]) -> dict[str, Any]:
        if not articles:
            return {"status": "OK", "response": {"log": []}}
        return self._post(
            "/api/productSet/remove/",
            {"token": self.token(), "articles": articles},
        )


def import_results(response: dict[str, Any]) -> dict[str, tuple[bool, str]]:
    status = str(response.get("status", "")).upper()
    nested = response.get("response")
    log_items = nested.get("log", []) if isinstance(nested, dict) else []
    results: dict[str, tuple[bool, str]] = {}
    if isinstance(log_items, list):
        for entry in log_items:
            if not isinstance(entry, dict):
                continue
            article = normalize(entry.get("article"))
            info = entry.get("info", [])
            messages = []
            codes = []
            if isinstance(info, list):
                for item in info:
                    if isinstance(item, dict):
                        codes.append(item.get("code"))
                        messages.append(normalize(item.get("message")))
            success = 0 in codes or (status == "OK" and not codes)
            results[article] = (success, "; ".join(message for message in messages if message))
    return results


def remove_results(response: dict[str, Any]) -> dict[str, tuple[bool, str]]:
    nested = response.get("response")
    log_items = nested.get("log", []) if isinstance(nested, dict) else []
    results: dict[str, tuple[bool, str]] = {}
    if not isinstance(log_items, list):
        return results
    for entry in log_items:
        if not isinstance(entry, dict):
            continue
        article = normalize(entry.get("article"))
        if article:
            results[article] = (bool(entry.get("deleted")), normalize(entry.get("message")))
    return results


def build_state_excel(entries: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Наборы"
    headers = ["Артикул набору", "Артикули відображення товарів", "Кінцева ціна", "Видалити (Так)"]
    worksheet.append(headers)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:D1"
    for column, width in {"A": 28, "B": 52, "C": 16, "D": 14}.items():
        worksheet.column_dimensions[column].width = width
    header_fill = PatternFill("solid", fgColor="166534")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for entry in entries:
        worksheet.append([
            entry.get("article", ""),
            "; ".join(entry.get("display_articles", [])),
            entry.get("discounted_price", ""),
            "",
        ])
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1).number_format = "@"
        worksheet.cell(row=row, column=2).number_format = "@"
        worksheet.cell(row=row, column=3).number_format = "0.00"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class StateStore:
    def __init__(self, path: Path):
        self.path = path
        self.sets: dict[str, dict[str, Any]] = {}
        self.history: list[dict[str, Any]] = []
        self.load()

    def load(self) -> None:
        if not self.path.exists():
            return
        with self.path.open("r", encoding="utf-8-sig") as file:
            data = json.load(file)
        if not isinstance(data, dict):
            raise HoroshopSetsError("Файл реестра наборов имеет неверный формат.")
        if isinstance(data.get("sets"), dict):
            self.sets = {
                normalize(article): value
                for article, value in data["sets"].items()
                if normalize(article) and isinstance(value, dict)
            }
        if isinstance(data.get("history"), list):
            self.history = [item for item in data["history"] if isinstance(item, dict)][-1000:]

    def _backup_current_file(self) -> None:
        if not self.path.exists():
            return
        backup_dir = self.path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
        backup = backup_dir / f"{self.path.stem}-{stamp}{self.path.suffix}"
        shutil.copy2(self.path, backup)
        backups = sorted(backup_dir.glob(f"{self.path.stem}-*{self.path.suffix}"))
        for old_backup in backups[:-50]:
            old_backup.unlink(missing_ok=True)

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._backup_current_file()
        temp = self.path.with_suffix(f"{self.path.suffix}.tmp")
        with temp.open("w", encoding="utf-8") as file:
            json.dump(
                {
                    "version": 2,
                    "updated_at": utc_now(),
                    "sets": self.sets,
                    "history": self.history[-1000:],
                },
                file,
                ensure_ascii=False,
                indent=2,
            )
        os.replace(temp, self.path)

    def contains(self, article: str) -> bool:
        return normalize(article) in self.sets

    def get(self, article: str) -> dict[str, Any] | None:
        return self.sets.get(normalize(article))

    def _append_history(self, action: str, article: str, status: str, message: str, snapshot: dict[str, Any] | None = None) -> None:
        event: dict[str, Any] = {
            "action": action,
            "article": article,
            "status": status,
            "message": message,
            "at": utc_now(),
        }
        if snapshot is not None:
            event["snapshot"] = snapshot
        self.history.append(event)
        self.history = self.history[-1000:]

    def record(self, item: PlanItem, status: str, message: str, source: str = "api") -> None:
        existing = self.get(item.article) or {}
        now = utc_now()
        entry = {
            "article": item.article,
            "display_articles": list(item.display_articles),
            "products": list(item.products),
            "discounted_price": str(item.discounted_price or ""),
            "title": SET_TITLE,
            "currency": item.currency or existing.get("currency", "UAH"),
            "enabled": item.enabled,
            "sort_order": item.sort_order,
            "discount_percent": item.discount_percent,
            "status": status,
            "message": message,
            "source": source,
            "created_at": existing.get("created_at", now),
            "updated_at": now,
        }
        self.sets[item.article] = entry
        self._append_history(item.action, item.article, status, message, entry)

    def record_failed_attempt(self, item: PlanItem, message: str) -> None:
        self._append_history(item.action, item.article, "error", message)

    def record_deletion(self, article: str, status: str, message: str) -> None:
        existing = self.sets.get(article)
        self._append_history("delete", article, status, message, existing)
        if status == "deleted":
            self.sets.pop(article, None)
        elif existing is not None:
            existing["status"] = status
            existing["message"] = message
            existing["updated_at"] = utc_now()

    def snapshot(self) -> list[dict[str, Any]]:
        return sorted(
            self.sets.values(),
            key=lambda item: str(item.get("updated_at", "")),
            reverse=True,
        )
