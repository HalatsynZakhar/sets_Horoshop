from __future__ import annotations

import io
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from horoshop_sets import (
    CatalogIndex,
    PlanItem,
    SetRow,
    Settings,
    StateStore,
    build_excel_template,
    build_state_excel,
    import_payload,
    import_results,
    load_settings,
    parse_excel_sets,
    prepare_plan,
)


class HoroshopSetsTests(unittest.TestCase):
    def excel_bytes(self, rows: list[tuple[object, object, object]]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        for row in rows:
            worksheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_template_has_expected_columns(self) -> None:
        workbook = load_workbook(io.BytesIO(build_excel_template()), read_only=True)
        self.assertEqual(
            [cell.value for cell in next(workbook.active.iter_rows(max_row=1))],
            [
                "Артикул набору (необов'язково)",
                "Артикули відображення товарів",
                "Кінцева ціна",
            ],
        )
        self.assertEqual(workbook.active["A2"].number_format, "@")
        self.assertEqual(workbook.active.title, "Набори")
        self.assertEqual(workbook["Інструкція"]["A1"].value, "Шаблон для створення та оновлення наборів")
        self.assertIn("Артикул набору необов'язковий", workbook["Інструкція"]["A4"].value)
        workbook.close()

    def test_excel_parses_price_and_articles(self) -> None:
        rows = parse_excel_sets(
            self.excel_bytes(
                [
                    ("Артикул набора", "Артикулы отображения товаров", "Цена набора"),
                    ("SET-1", "show-a; SHOW-B", "199,50"),
                ]
            )
        )
        self.assertEqual(rows, [SetRow("SET-1", ("show-a", "SHOW-B"), Decimal("199.50"), 2)])

    def test_excel_builds_article_when_it_is_blank(self) -> None:
        rows = parse_excel_sets(
            self.excel_bytes(
                [
                    ("Артикул набору (необов'язково)", "Артикули відображення товарів", "Кінцева ціна"),
                    (None, "show-a; SHOW-B", "199,50"),
                ]
            )
        )
        self.assertEqual(rows[0].article, "show-a.SHOW-B")

    def test_excel_supports_delete_action_without_price_or_products(self) -> None:
        rows = parse_excel_sets(
            self.excel_bytes(
                [
                    ("Артикул набору", "Артикули відображення товарів", "Кінцева ціна", "Видалити (Так)"),
                    ("SET-1", None, None, "Так"),
                ]
            )
        )
        self.assertEqual(rows[0].article, "SET-1")
        self.assertEqual(rows[0].action, "delete")
        self.assertIsNone(rows[0].discounted_price)

    def test_excel_treats_blank_delete_marker_as_upsert(self) -> None:
        rows = parse_excel_sets(
            self.excel_bytes(
                [
                    ("Артикул набору", "Артикули відображення товарів", "Кінцева ціна", "Видалити (Так)"),
                    ("SET-1", "A; B", 100, ""),
                ]
            )
        )
        self.assertEqual(rows[0].action, "upsert")
        self.assertEqual(rows[0].display_articles, ("A", "B"))

    def test_registry_export_has_explicit_delete_marker(self) -> None:
        workbook = load_workbook(
            io.BytesIO(build_state_excel([{"article": "SET-1", "display_articles": ["A", "B"], "discounted_price": "100"}])),
            read_only=True,
        )
        worksheet = workbook.active
        self.assertEqual(worksheet["D1"].value, "Видалити (Так)")
        self.assertIsNone(worksheet["D2"].value)
        workbook.close()

    def test_catalog_resolution_prefers_exact_then_case_insensitive(self) -> None:
        catalog = CatalogIndex.from_raw(
            [
                {"article": "REAL-A", "article_for_display": "Display-A"},
                {"article": "REAL-B", "article_for_display": "DISPLAY-B"},
            ]
        )
        plan = prepare_plan(
            [SetRow("SET-1", ("Display-A", "display-b"), Decimal("12.00"), 2)],
            catalog,
        )
        self.assertTrue(plan[0].ready)
        self.assertEqual(plan[0].products, ("REAL-A", "REAL-B"))

    def test_catalog_index_accepts_raw_api_export(self) -> None:
        catalog = CatalogIndex.from_raw(
            [{"article": "REAL-A", "article_for_display": "DISPLAY-A"}]
        )
        article, error = catalog.resolve_display_article("DISPLAY-A")
        self.assertEqual(article, "REAL-A")
        self.assertEqual(error, "")

    def test_ambiguous_display_article_and_product_article_collision_are_blocked(self) -> None:
        catalog = CatalogIndex.from_raw(
            [
                {"article": "PRODUCT-1", "article_for_display": "Same"},
                {"article": "PRODUCT-2", "article_for_display": "same"},
            ]
        )
        plan = prepare_plan(
            [
                SetRow("SET-1", ("sAmE", "other"), Decimal("12.00"), 2),
                SetRow("PRODUCT-1", ("Same", "other"), Decimal("12.00"), 3),
            ],
            catalog,
        )
        self.assertIn("не є унікальним", plan[0].error)
        self.assertIn("збігається", plan[1].error)

    def test_state_records_composition(self) -> None:
        item = PlanItem("SET-1", ("A", "B"), Decimal("10.00"), ("REAL-A", "REAL-B"), 2)
        with tempfile.TemporaryDirectory() as directory:
            store = StateStore(Path(directory) / "state.json")
            store.record(item, "synced", "Imported")
            store.save()
            reloaded = StateStore(Path(directory) / "state.json")
        self.assertEqual(reloaded.snapshot()[0]["products"], ["REAL-A", "REAL-B"])
        self.assertEqual(reloaded.snapshot()[0]["status"], "synced")

    def test_deletion_removes_active_set_and_keeps_history(self) -> None:
        item = PlanItem("SET-1", ("A", "B"), Decimal("10.00"), ("REAL-A", "REAL-B"), 2)
        with tempfile.TemporaryDirectory() as directory:
            store = StateStore(Path(directory) / "state.json")
            store.record(item, "synced", "Imported")
            store.record_deletion("SET-1", "deleted", "Deleted")
            store.save()
            reloaded = StateStore(Path(directory) / "state.json")
        self.assertEqual(reloaded.snapshot(), [])
        self.assertEqual(reloaded.history[-1]["action"], "delete")

    def test_payload_and_api_log_use_discounted_price_only(self) -> None:
        item = PlanItem("SET-1", ("A", "B"), Decimal("10.50"), ("REAL-A", "REAL-B"), 2)
        settings = Settings(
            domain="https://shop.example.com",
            host="0.0.0.0",
            port=8093,
            currency="UAH",
            title="Разом дешевше",
            batch_size=50,
            request_timeout_seconds=60,
            state_file=Path("state.json"),
        )
        payload = import_payload([item], settings)
        self.assertEqual(payload[0]["discountedPrice"], 10.5)
        self.assertEqual(payload[0]["title"], "Разом дешевше")
        self.assertNotIn("initialPrice", payload[0])
        self.assertEqual(
            import_results(
                {"status": "WARNING", "response": {"log": [{"article": "SET-1", "info": [{"code": 0, "message": "Imported"}]}]}}
            )["SET-1"],
            (True, "Imported"),
        )

    def test_settings_accept_utf8_bom_from_windows_notepad(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config_path = Path(directory) / "config.json"
            config_path.write_text(
                '{"server":{"port":8093},"horoshop":{"domain":"https://shop.example.com","title":"Вместе дешевле"},"logging":{"public_log_path":"public-logs","public_log_name":"visible.log"}}',
                encoding="utf-8-sig",
            )
            settings = load_settings(config_path)
        self.assertEqual(settings.domain, "https://shop.example.com")
        self.assertEqual(settings.title, "Разом дешевше")
        self.assertEqual(settings.public_log_file, config_path.parent / "public-logs" / "visible.log")

    def test_settings_repair_unescaped_windows_public_log_path(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config_path = Path(directory) / "config.json"
            config_path.write_text(
                r'{"horoshop":{"domain":"https://shop.example.com"},"logging":{"public_log_path":"C:\ShareFiles\public","public_log_name":"visible.log"}}',
                encoding="utf-8",
            )
            settings = load_settings(config_path)
            repaired_config = config_path.read_text(encoding="utf-8")
        self.assertEqual(settings.public_log_file, Path(r"C:\ShareFiles\public") / "visible.log")
        self.assertIn(r'"public_log_path":"C:\\ShareFiles\\public"', repaired_config)


if __name__ == "__main__":
    unittest.main()
